"""
PDF Studio - Módulo de Conversão PDF para Excel
Funcionalidades para extração de tabelas e conversão para Excel
"""

import os
import pdfplumber
import tabula
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re
from .pdf_utils import generate_unique_filename, cleanup_file

class PDFConverter:
    def __init__(self, upload_folder, output_folder):
        self.upload_folder = upload_folder
        self.output_folder = output_folder
    
    def parse_text_to_table(self, text):
        """Parse text content to extract structured data as table"""
        lines = text.strip().split('\n')
        table_data = []
        
        # Try to detect patterns like the example shown
        # Pattern: ID + Description + Unit + Value + Date
        # Followed by: Company + Unit + Value + Value + Percentages
        
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            
            # Skip empty lines
            if not line:
                i += 1
                continue
                
            # Check if this line starts with a pattern like "10 13866" (ID + number)
            if re.match(r'^\d+\s+\d+', line):
                # This looks like a main item line
                parts = line.split()
                
                if len(parts) >= 5:
                    # Extract ID, description, unit, value, date
                    item_id = f"{parts[0]} {parts[1]}"
                    
                    # Find description (everything between ID and the last 3 parts)
                    desc_parts = parts[2:-3]
                    description = ' '.join(desc_parts)
                    
                    # Last 3 parts should be unit, value, date
                    unit = parts[-3]
                    value = parts[-2]
                    date = parts[-1]
                    
                    # Look for the next line (company info)
                    company_info = ""
                    if i + 1 < len(lines):
                        next_line = lines[i + 1].strip()
                        if next_line and not re.match(r'^\d+\s+\d+', next_line):
                            company_info = next_line
                            i += 1  # Skip the next line as we processed it
                    
                    # Create table row
                    if company_info:
                        # Split company info into parts
                        company_parts = company_info.split()
                        if len(company_parts) >= 5:
                            company_name = ' '.join(company_parts[:-4])  # Everything except last 4 parts
                            company_unit = company_parts[-4]
                            company_value1 = company_parts[-3]
                            company_value2 = company_parts[-2]
                            percentages = company_parts[-1]
                            
                            table_data.append([
                                item_id,
                                description,
                                unit,
                                value,
                                date,
                                company_name,
                                company_unit,
                                company_value1,
                                company_value2,
                                percentages
                            ])
                        else:
                            # Fallback: just add company info as one field
                            table_data.append([
                                item_id,
                                description,
                                unit,
                                value,
                                date,
                                company_info,
                                "", "", "", ""
                            ])
                    else:
                        # No company info, just main item
                        table_data.append([
                            item_id,
                            description,
                            unit,
                            value,
                            date,
                            "", "", "", "", ""
                        ])
            else:
                # If line doesn't match pattern, try to extract as generic data
                # Split by multiple spaces or tabs
                parts = re.split(r'\s{2,}|\t', line)
                if len(parts) >= 3:  # At least 3 columns
                    cleaned_parts = [p.strip() for p in parts if p.strip()]
                    if len(cleaned_parts) >= 3:
                        table_data.append(cleaned_parts)
            
            i += 1
        
        # If we found data with specific pattern, add headers
        if table_data:
            # Check if we used the specific pattern (10 columns) or generic
            if len(table_data) > 0 and len(table_data[0]) == 10:
                headers = [
                    "ID",
                    "Descrição",
                    "Unidade",
                    "Valor",
                    "Data",
                    "Empresa",
                    "Unidade Empresa",
                    "Valor 1",
                    "Valor 2",
                    "Percentagens"
                ]
                return [headers] + table_data
            else:
                # Generic data - use generic headers
                max_cols = max(len(row) for row in table_data) if table_data else 0
                if max_cols > 0:
                    headers = [f"Coluna {i+1}" for i in range(max_cols)]
                    # Pad rows
                    padded_data = []
                    for row in table_data:
                        padded_row = row + [""] * (max_cols - len(row))
                        padded_data.append(padded_row)
                    return [headers] + padded_data
        
        # If no structured data found, try a more generic approach
        return self.parse_generic_text_to_table(text)
    
    def parse_generic_text_to_table(self, text):
        """Generic text parsing for any structured data"""
        lines = text.strip().split('\n')
        table_data = []
        
        # Look for any lines that contain multiple data points separated by spaces
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # Split by multiple spaces to find data columns
            parts = re.split(r'\s{2,}', line)  # Split by 2 or more spaces
            
            if len(parts) >= 2:  # At least 2 columns of data (reduced from 3)
                # Clean up each part
                cleaned_parts = [part.strip() for part in parts if part.strip()]
                if len(cleaned_parts) >= 2:  # At least 2 columns
                    table_data.append(cleaned_parts)
            else:
                # Try splitting by single spaces if multiple spaces didn't work
                parts = line.split()
                if len(parts) >= 3:  # At least 3 words/columns
                    table_data.append(parts)
        
        # If we found data, create headers
        if table_data:
            max_cols = max(len(row) for row in table_data) if table_data else 0
            if max_cols > 0:
                headers = [f"Coluna {i+1}" for i in range(max_cols)]
                
                # Pad rows to have same number of columns
                padded_data = []
                for row in table_data:
                    padded_row = row + [""] * (max_cols - len(row))
                    padded_data.append(padded_row)
                
                return [headers] + padded_data
        
        return None
    
    def extract_tables_pdfplumber(self, pdf_path):
        """Extract tables using pdfplumber - primary method"""
        tables = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    print(f"\n=== Processing Page {page_num + 1} ===")
                    
                    # Strategy 1: Try with default settings first (most reliable)
                    page_tables = page.extract_tables()
                    print(f"Default extraction: Found {len(page_tables) if page_tables else 0} tables")
                    
                    # Strategy 2: If default finds tables but they seem incomplete, try with lines strategy
                    if page_tables:
                        total_rows = sum(len(t) for t in page_tables if t)
                        print(f"Total rows found: {total_rows}")
                        
                        # If very few rows, try alternative strategies
                        if total_rows < 5:
                            print("Few rows detected, trying alternative extraction...")
                            # Try with explicit line detection
                            alt_tables = page.extract_tables(table_settings={
                                "vertical_strategy": "lines_strict",
                                "horizontal_strategy": "lines_strict",
                                "snap_tolerance": 5,
                                "join_tolerance": 5,
                            })
                            if alt_tables:
                                alt_total_rows = sum(len(t) for t in alt_tables if t)
                                print(f"Alternative extraction found {alt_total_rows} rows")
                                if alt_total_rows > total_rows:
                                    page_tables = alt_tables
                                    print("Using alternative extraction (more rows)")
                    
                    # Strategy 3: If still no tables or very few, try text-based
                    if not page_tables or (page_tables and sum(len(t) for t in page_tables if t) < 3):
                        print("Trying text-based extraction...")
                        text_tables = page.extract_tables(table_settings={
                            "vertical_strategy": "text",
                            "horizontal_strategy": "text",
                        })
                        if text_tables:
                            text_total_rows = sum(len(t) for t in text_tables if t)
                            print(f"Text-based extraction found {text_total_rows} rows")
                            if not page_tables or text_total_rows > sum(len(t) for t in page_tables if t):
                                page_tables = text_tables
                                print("Using text-based extraction")
                    
                    structured_tables_found = False
                    
                    if page_tables:
                        print(f"Processing {len(page_tables)} table(s)...")
                        
                        # If multiple small tables, try to merge them (might be one table split)
                        if len(page_tables) > 1:
                            total_rows_all = sum(len(t) for t in page_tables if t)
                            print(f"Multiple tables detected ({len(page_tables)}), total rows: {total_rows_all}")
                            
                            # Check if tables have similar column structure (likely parts of same table)
                            if total_rows_all > 0:
                                first_table_cols = len(page_tables[0][0]) if page_tables[0] and page_tables[0][0] else 0
                                similar_cols = all(
                                    len(t[0]) == first_table_cols 
                                    for t in page_tables 
                                    if t and t[0] and len(t[0]) > 0
                                ) if first_table_cols > 0 else False
                                
                                if similar_cols and first_table_cols >= 5:  # Likely same table split
                                    print("Tables appear to have same structure, merging...")
                                    merged_table = []
                                    max_cols = 0
                                    
                                    for table in page_tables:
                                        if table:
                                            for row in table:
                                                if row:
                                                    cleaned_row = [str(cell).strip() if cell else '' for cell in row]
                                                    # ALWAYS add row - don't filter!
                                                    merged_table.append(cleaned_row)
                                                    max_cols = max(max_cols, len(cleaned_row))
                                    
                                    if merged_table:
                                        # Normalize columns
                                        for i, row in enumerate(merged_table):
                                            while len(row) < max_cols:
                                                row.append('')
                                            merged_table[i] = row
                                        
                                        print(f"Merged table: {len(merged_table)} rows, {max_cols} columns")
                                        structured_tables_found = True
                                        tables.append({
                                            'page': page_num + 1,
                                            'table': 1,
                                            'data': merged_table
                                        })
                        
                        # If not merged, process tables individually
                        if not structured_tables_found:
                            for table_num, table in enumerate(page_tables):
                                if table and len(table) > 0:
                                    # Clean the table data
                                    cleaned_table = []
                                    max_cols = 0
                                    
                                    for row_idx, row in enumerate(table):
                                        if row:  # Row exists
                                            # Clean all cells
                                            cleaned_row = []
                                            for cell in row:
                                                if cell:
                                                    cleaned_cell = str(cell).strip()
                                                    cleaned_row.append(cleaned_cell)
                                                else:
                                                    cleaned_row.append('')
                                            
                                            # ALWAYS add row - don't filter!
                                            cleaned_table.append(cleaned_row)
                                            max_cols = max(max_cols, len(cleaned_row))
                                    
                                    # Normalize all rows to have the same number of columns
                                    if cleaned_table:
                                        for i, row in enumerate(cleaned_table):
                                            while len(row) < max_cols:
                                                row.append('')
                                            cleaned_table[i] = row
                                        
                                        structured_tables_found = True
                                        print(f"  Table {table_num + 1}: {len(cleaned_table)} rows, {max_cols} columns")
                                        if len(cleaned_table) > 0:
                                            print(f"    First row: {cleaned_table[0][:5]}...")  # Show first 5 columns
                                            if len(cleaned_table) > 1:
                                                print(f"    Last row: {cleaned_table[-1][:5]}...")
                                        
                                        tables.append({
                                            'page': page_num + 1,
                                            'table': table_num + 1,
                                            'data': cleaned_table
                                        })
                    
                    # If no structured tables found or too few rows, try text extraction as fallback
                    if not structured_tables_found or (structured_tables_found and len(tables) > 0 and len(tables[-1]['data']) < 3):
                        print("Trying full text extraction as fallback...")
                        text = page.extract_text()
                        if text:
                            print(f"Extracted text length: {len(text)} characters")
                            parsed_data = self.parse_text_to_table(text)
                            if parsed_data and len(parsed_data) > 1:
                                print(f"Text parsing found {len(parsed_data)} rows")
                                # Only add if we don't have tables or if text parsing found more rows
                                if not structured_tables_found or (parsed_data and len(parsed_data) > len(tables[-1]['data']) if tables else False):
                                    print("Using text-parsed data")
                                    tables.append({
                                        'page': page_num + 1,
                                        'table': len(page_tables) + 1 if page_tables else 1,
                                        'data': parsed_data
                                    })
                    
                    print(f"Page {page_num + 1} complete: {len([t for t in tables if t['page'] == page_num + 1])} table(s) added")
                    
        except Exception as e:
            print(f"Error with pdfplumber: {e}")
            import traceback
            traceback.print_exc()
            return None
        return tables
    
    def extract_tables_tabula(self, pdf_path):
        """Extract tables using tabula-py - fallback method"""
        tables = []
        try:
            # Try to extract all tables from all pages
            dfs = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)
            
            for page_num, df in enumerate(dfs):
                if not df.empty:
                    # Convert DataFrame to list of lists
                    table_data = [df.columns.tolist()] + df.values.tolist()
                    
                    # Clean the data
                    cleaned_table = []
                    for row in table_data:
                        cleaned_row = [str(cell).strip() if pd.notna(cell) else '' for cell in row]
                        cleaned_table.append(cleaned_row)
                    
                    tables.append({
                        'page': page_num + 1,
                        'table': 1,
                        'data': cleaned_table
                    })
        except Exception as e:
            print(f"Error with tabula: {e}")
            # If tabula fails due to Java issues, try text extraction as fallback
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    for page_num, page in enumerate(pdf.pages):
                        text = page.extract_text()
                        if text:
                            parsed_data = self.parse_text_to_table(text)
                            if parsed_data:
                                tables.append({
                                    'page': page_num + 1,
                                    'table': 1,
                                    'data': parsed_data
                                })
            except Exception as e2:
                print(f"Error with text extraction fallback: {e2}")
                return None
        return tables
    
    def create_excel_file(self, tables, output_path):
        """Create Excel file with proper formatting"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        for table_info in tables:
            table_data = table_info['data']
            page_num = table_info['page']
            table_num = table_info['table']
            
            if not table_data or len(table_data) == 0:
                print(f"Skipping empty table on page {page_num}, table {table_num}")
                continue
            
            # Create sheet name
            if len(tables) == 1:
                sheet_name = f"Page_{page_num}"
            else:
                sheet_name = f"Page_{page_num}_Table_{table_num}"
            
            # Ensure sheet name is valid (max 31 chars, no invalid chars)
            sheet_name = sheet_name[:31]
            sheet_name = sheet_name.replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
            
            # If sheet name already exists, append number
            original_name = sheet_name
            counter = 1
            while sheet_name in [ws.title for ws in wb.worksheets]:
                sheet_name = f"{original_name}_{counter}"[:31]
                counter += 1
            
            ws = wb.create_sheet(title=sheet_name)
            
            print(f"Creating sheet '{sheet_name}' with {len(table_data)} rows")
            
            # Add data to worksheet
            for row_idx, row in enumerate(table_data, 1):
                for col_idx, cell_value in enumerate(row, 1):
                    # Handle None values
                    if cell_value is None:
                        cell_value = ''
                    
                    # Convert to string and clean
                    cell_value = str(cell_value).strip() if cell_value else ''
                    
                    try:
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    except Exception as e:
                        print(f"Error writing cell ({row_idx}, {col_idx}): {e}")
                        ws.cell(row=row_idx, column=col_idx, value='')
            
            # Format header row (first row)
            if table_data:
                header_row = 1
                max_cols = max(len(row) for row in table_data) if table_data else 0
                
                for col_idx in range(1, max_cols + 1):
                    cell = ws.cell(row=header_row, column=col_idx)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            max_cols = max(len(row) for row in table_data) if table_data else 0
            for col_idx in range(1, max_cols + 1):
                max_length = 0
                column_letter = ws.cell(row=1, column=col_idx).column_letter
                
                for row_idx in range(1, len(table_data) + 1):
                    try:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # Set width with reasonable limits
                adjusted_width = min(max(max_length + 2, 10), 50)  # Min 10, Max 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
        
        # Save Excel file
        wb.save(output_path)
        print(f"Excel file saved: {output_path}")
    
    def convert_pdf_to_excel(self, pdf_path):
        """
        Converte PDF para Excel
        
        Args:
            pdf_path: Caminho do PDF
            
        Returns:
            tuple: (success, excel_path, error_message)
        """
        try:
            # Gerar nome único para arquivo Excel
            excel_filename, file_id = generate_unique_filename("converted.xlsx")
            excel_path = os.path.join(self.output_folder, excel_filename)
            
            # Extract tables using pdfplumber first
            print(f"Extracting tables from: {pdf_path}")
            tables = self.extract_tables_pdfplumber(pdf_path)
            print(f"Found {len(tables) if tables else 0} tables with pdfplumber")
            
            # If pdfplumber fails or returns empty, try tabula
            if not tables:
                print("Trying tabula-py as fallback...")
                tables = self.extract_tables_tabula(pdf_path)
                print(f"Found {len(tables) if tables else 0} tables with tabula")
            
            if not tables:
                return False, None, "Nenhuma tabela encontrada no PDF. Certifique-se de que o PDF contém dados tabulares."
            
            # Debug: print table info
            total_rows = 0
            for table in tables:
                rows = len(table.get('data', []))
                total_rows += rows
                print(f"Table on page {table['page']}: {rows} rows")
            
            print(f"Total rows to export: {total_rows}")
            
            # Create Excel file
            self.create_excel_file(tables, excel_path)
            
            return True, excel_path, None
            
        except Exception as e:
            import traceback
            error_msg = f"Erro ao converter PDF: {str(e)}\n{traceback.format_exc()}"
            print(error_msg)
            return False, None, error_msg

